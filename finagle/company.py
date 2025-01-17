import pandas as pd
import xlrd
from string import ascii_letters
import datetime
import numpy as np
from scipy import interpolate
import logging
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

pd.set_option('mode.chained_assignment', None)


class company:
    '''A class to model the financials of a publicly traded company.

    It has a number of method types which should be run in order:
    __init__(), will load the data into the appropriate attributes and
    fcf_from_X(), there are several of these methods. They are all meant to be
    used for calculating fcf
    fcf_to_X(), these are methods which are used to model how the FCF is to be
    used: paying down debt, buying back shares value(), this method is use for
    calculating a DCF from the cashflows, This shold be run only once other
    methods described have been run display_fin(), this method is used to
    process the financials. It should be used once all modelling is completed
    
    the financials dictonary should have the following keys:
    date = date of the last reporting quarter, list size 1
    ebitda = earnings before, interest, tax, depreciation or amortization, list size 1 or years
    capex = capital expenditure, list size 1 or years
    dwc = changes in non-cash workking capital, list size years
    tax = ttm reported taxes, list size 1
    da = ttm depreciation and amortization, list size 1
    debt = short and long term debt, list of size years
    cash = excess cash, short term investments on balancesheet not used for working capital, list size 1
    nol = net operating loss, i.e. tax losses subtracted from ebit for the purpose of calculating taxes, list size 1
    noa = other non-operating assets, examples: land, properties, other even business's not included in the cashflows, list size 1.

    Args:
        financials (dict): python dictionary of financial input data.
        Each value can be a single value or a list. First value in the 0 index
        location shoulds be the value which corresponds to the trailing twelve
        months (ttm) actual keys required by the object will depend on which
        subsequent methods are used. See methods doc string for specific
        guidance.
        ticker (str): ticker symbol of the company
        re (float): cost of equity
        rd (float): cost of debt
        t (float): marginal tax rate
        te (float): effective tax rate for year 1
        shares (int): shares outstanding. Should include all classes if
        multiple classes exist
        price (float): current share price
        gt (float): terminal growth, required for model closure
        roict (float): terminal return on invested capital, used for
        calculating depreciation
        year (int): final forecast year. Year before the terminal year.
        fcfe (list,None): free cash flow to equity
        fcff (list,None): free cash flow to firm
        roict (float): return on invested capital in the terminal year, 15%
        default, mostly creates impact via taxes (by calculating depreciation)
        since gt and capex are explicitly specified
        dividend (float,list)): current dividend policy

    Attributes:
        logfile (str): name of logfile
        ticker (str): ticker symbol of the company
        gt (float): terminal growth, required for model closure
        re (float): cost of equity
        rd (float): cost of debt
        t (float): marginal tax rate
        te (float): effective tax rate for year 1
        roict (float): terminal return on invested capital, used for
        calculating depreciation
        shares (int): shares outstanding. Should include all classes if
        multiple classes exist
        price (float): current share price
        year (int): final forecast year. Year before the terminal year.
        years (list): list of years, starting at 0 (current year)
        now (date): todays date
        buybacks (bool):
        dividend (list): current dividend policy
    '''

    def __init__(self, financials=None, ticker=None, re=None, rd=None, t=None, te=None,
                 shares=1, price=0, gt=0, fcfe=None, fcff=None, fcf=None,
                 roict=0.15, year=6, dividend=0):

        # setup logging
        self.logfile = ticker + '.log'
        logging.basicConfig(filemode='w', level=logging.INFO,
                            format='%(asctime)s %(levelname)s:%(message)s')
        logging.FileHandler(filename=self.logfile, mode='w')
        logging.info(ticker)
        
        # read in various attributes
        self.ticker = ticker
        self.gt = gt
        self.re = re
        self.rd = rd
        self.t = t
        self.te = te
        self.roict = roict
        self.shares = shares
        self.price = price
        self.year = year
        self.years = list(range(year+1))
        self.now = datetime.date.today()
        self.buybacks = False

        if isinstance(dividend, list):
            self.dividend = dividend
        elif isinstance(dividend, float):
            self.dividend = [dividend]
        elif isinstance(dividend, int):
            self.dividend = [dividend]

        # create financial dataframe
        if financials is None:
            pass
        else:
            self.load_financials(financials=financials)

            # initialize the FCF columns, if you want to do a DCF directly
            # without calculating from financials, this requires a financial
            # dict with a date for the ttm year
            self.fin['fcfe'] = fcfe
            self.fin['fcff'] = fcff
            self.fin['fcf'] = fcf
            # buybacks will always be 0 here, initially ignore dividend policy
            # and distribute all cash
            self.fin['dividend'] = (
                self.fin['fcfe']-self.fin['buybacks'])/self.fin['shares']

    def __stream(self, sf, st):
        '''create a periodic stream of values based of yearly forecasts and
        terminal values

        Args:
            sf: forecasted stream values
            st: stream value in terminal year

        Returns:
            values:
        '''

        if np.isnan(st) is True:
            logging.error('Terminal value is a NaN, cannot create forecast')

        if isinstance(sf, list):
            length = np.count_nonzero(~np.isnan(sf))
        elif isinstance(sf, float):
            length = 1
            sf = [sf]
        else:
            length = np.count_nonzero(~np.isnan(sf))
            sf = list(sf.iloc[0:length])

        x = list(range(length))
        x.append(self.year)
        s = sf+[st]
        f = interpolate.interp1d(x, s)
        values = f(self.years)
        return values

    def __datacheck(self):
        '''check for all required columns, check sufficient number of years,
        create flags for which fcf_from_x functions may be used
        '''

        self.data_for_earnings = False
        self.data_for_ebitda = False

        if 'ebitda' in self.fin.columns:
            if self.year < len(self.fin.ebitda)-1:
                self.year = len(self.fin.ebitda)-1
                self.years = list(range(self.year+1))
                logging.warning("Warning: length of EBITDA forecast appear larger then the 'year' parameter used at initialization")

            from_ebitda_columns = ['revenue','price', 'tax', 'interest', 'capex', 'noa',
                                   'nol', 'ebitda', 'shares', 'dwc', 'debt',
                                   'cash', 'da', 'MnA', 'buybacks', 'cashBS',
                                   'sbc']
            from_ebitda_forecasts = ['ebitda', 'capex', 'dwc', 'debt', 'sbc']

            check_columns = set(from_ebitda_columns) == set(
                list(self.fin.columns.values))  # check all columns are present
            check_forecasts = np.count_nonzero(
                np.isnan(self.fin[from_ebitda_forecasts])) == 0
            # check for nan's
            self.data_for_ebitda = check_columns and check_forecasts

        if 'e' in self.fin.columns:
            self.data_for_earnings = True

        if self.data_for_ebitda is True:
            logging.info('datacheck complete: financial dataset appears complete for fcf_from_ebitda')

        if self.data_for_earnings is True:
            logging.info('datacheck complete: financial dataset appears complete for fcf_from_earnings')

    def __pv(self, cfs, g, r, cft=None):
        '''calculate the present value of future cash flows. To be even more
        clear, any flows from the current year are not included in the present
        value calculation - only future years the last cf in cfs is the year
        before the terminal year

        Args:
            cfs: array of cash flows
            g:
            r:
            cft: termal cash flow

        Returns:
            value:
        '''
        if isinstance(r, float):
            r = [r for cf in cfs]
        else:
            r = list(r)

        if cft is None:
            value = [cfs.iloc[-1]*(1+g)/(r[-1]-g)]
        else:
            value = [cft/(r[-1]-g)]
        r = r[::-1]
        for i, cf in enumerate(cfs[:0:-1], 1):
            value.insert(0, (cf+value[0])/(1+r[i]))
        return value

    def __wacc(self):
        '''return weighted average cost of capital

        Returns:
            self.fin['wacc']:
        '''
        self.fin['EV'] = self.fin.debt + self.fin.equity
        self.fin['wacc'] = (self.fin.debt*self.rd*(1-self.t) +
                            self.fin.equity*self.re)/self.fin.EV
        return self.fin['wacc']

    def forecast_ebitda(self, ebitda_ttm, gf, financials=None,me=None,mc=None,gsnext=None):
        '''creates an ebitda forecast and populates the financials

        Args:
            ebitda_ttm: last year ebitda
            gf: ebitda growth rates
            financials:
            mc: ebitda contribution margin of the final year of forecast
            me: ebitda margin of the final year of forecast
            gsnext: sales growth forecast of the year following the last EBITDA growth rates forecast
            

        Return:
            ebitda:

        '''
        g = self.__stream(gf, self.gt)            
        ebitda = [ebitda_ttm]
        revenue = [np.nan]
        
        if mc and me and gsnext is not None:

            if isinstance(gf, list):
                length = np.count_nonzero(~np.isnan(gf))
            elif isinstance(gf, float):
                length = 1
                gf = [gf] 

            gfs = [0 for x in range(length)]
            gfs.append(gsnext)
            gs = self.__stream(gfs,self.gt)            

            for i in range(self.year):
                if i < length: 
                    revenue.append(None)
                    ebitda.append(ebitda[i]*(1+g[i]))
                else:
                    #print(me)
                    #print(ebitda[i]*mc/me*(gs[i])) #incremental ebitda
                    revenue.append(ebitda[i]/me*(1+gs[i])) #sales in the following year
                    ebitda.append(ebitda[i]*(1+mc/me*(gs[i])))
                    me=ebitda[i+1]/(ebitda[i]/me*(1+gs[i]))
        
        else:
            for i in range(self.year):
                revenue.append(0)
                ebitda.append(ebitda[i]*(1+g[i]))

        if financials is not None:
            # a dict it will populate the elements similar to a pointer
            financials['ebitda'] = ebitda
            financials['revenue'] = revenue

        return ebitda

    def forecast_capex(self, capex_f, financials):
        '''creates a capital expenditures (capex) forecast and populates the
        financials dictionary.

        This is done by proving a forecast in the form of a list. The length of
        the forecast can be for all years or just a subset. The balance of the
        years not provided are then forecast as a constant fraction of ebitda,
        using the last value in the capex_f list. The forecast is then used to
        populate the financials dictionary or returned as a list.

        Args:
            capex_f:
            financials:

        Returns:
            capex:
        '''
        if isinstance(capex_f, list):
            length = np.count_nonzero(~np.isnan(capex_f))
            capex = capex_f
        elif isinstance(capex_f, float):
            length = 1
            capex = [capex_f]
        elif isinstance(capex_f, int):
            length = 1
            capex = [capex_f]
        else:
            length = np.count_nonzero(~np.isnan(capex_f))
            capex = list(capex_f.iloc[0:length])

        # start scaling with ebitda where the forecast ends
        for i in range(length, self.year+1):
            capex.append(capex[length-1]/financials['ebitda']
                         [length-1]*financials['ebitda'][i])

        if financials is not None:
            # a dict it will populate the elements similar to a pointer
            financials['capex'] = capex

        return capex

    def forecast_sbc(self, sbc_f, financials, sbc_rate_t=None):
        '''creates an stock-based compensation (sbc) forecast and populates the
        financials dictionary.

        This is done by proving a forecast in the form of a list. The length of
        the forecast can be for all years or just a subset. The balance of the
        years not provided are then forecast using interpolation of the
        terminal fraction (sbc_rate_t) or at a constant rate using the last
        value in sbc_f. The forecast is then used to populate the financials
        dictionary or returned as a list.

        Args:
            sbc_f (float or list): current year sbc or forecast of multiple
            years
            sbc_rate_t (float): terminal sbc as a fraction of EBITDA, i.e. a
            number from 0-1, if None, then it scales from the last value in the
            forecast
            financials (dict): financials dictionary

        Returns:
            sbc: a list of the sbc.
        '''
        if isinstance(sbc_f, list):
            length = np.count_nonzero(~np.isnan(sbc_f))
            sbc = sbc_f
        elif isinstance(sbc_f, float):
            length = 1
            sbc = [sbc_f]
        elif isinstance(sbc_f, int):
            length = 1
            sbc = [sbc_f]
        else:
            length = np.count_nonzero(~np.isnan(sbc_f))
            sbc = list(sbc_f.iloc[0:length])

        if sbc_rate_t is None:
            sbc_rate_t = sbc[length-1]/financials['ebitda'][length-1]

        sbc_rate_f = []
        for i in range(length):
            sbc_rate_f.append(sbc[i]/financials['ebitda'][i])

        sbc_rate = self.__stream(sbc_rate_f, sbc_rate_t)
        #print(sbc_rate)
        for i in range(length, self.year+1):  
            # start scaling where the forecast ends
            sbc.append(sbc_rate[i]*financials['ebitda'][i])

        if financials is not None:
            # since this is a dict it will populate the elements similar to a
            # pointer
            financials['sbc'] = sbc

        return sbc

    def load_financials(self, financials):
        '''
        Args:

        Returns:
        '''
        financials['date'] = [datetime.datetime.strptime(
            financials['date'], '%Y-%m-%d')+datetime.timedelta(days=365*i) for i in range(self.year+1)]
        self.fin = pd.DataFrame({key: pd.Series(value)
                                for key, value in financials.items()})
        self.fin.set_index('date', inplace=True)
        self.fin['shares'] = self.shares
        self.fin['price'] = self.price
        self.fin['MnA'] = 0
        self.fin['buybacks'] = 0
        self.fin['cashBS'] = 0

        try:
            self.fin['cashBS'].iloc[0] = self.fin['cash'].iloc[0]
            self.cash0 = self.fin['cash'].iloc[0]
        except:
            logging.info('no cash key')

        self.__datacheck()
        logging.info('input data used for the forecast is:')
        logging.info(financials)
        logging.info('load_financials() method complete')

    def fcf_from_earnings(self, payout=1, gf=0, ROE=1):
        '''
        Args:
            payout = list of yearly payout forecasts, or most recent years
            ROE = perpetual return on equity
            gf = list of yearly growth forecast's

        Returns:
        '''

        if self.data_for_earnings is False:
            logging.error('financial dataset cannot be used for calculating FCF from earnings')

        g = self.__stream(gf, self.gt)

        for i in range(self.year):
            self.fin['e'].iloc[i+1] = self.fin['e'].iloc[i]*(1+g[i])

        payout_t = 1 - self.gt/ROE
        payouts = self.__stream(payout, payout_t)

        self.fin['fcfe'] = self.fin['e']*payouts
        logging.info('fcf_from_earnings() method complete')

    def fcf_from_ebitda(self):
        '''
        Args:
        Returns:
        '''

        if self.data_for_ebitda is False:
            logging.error('financial dataset cannot be used for calculating FCF from EBITDA')

        interest0 = self.fin['interest'].iloc[0]
        self.fin['interest'] = self.rd*self.fin.debt.shift(1)
        self.fin['interest'].iloc[0] = interest0
        
        # really complicated way to calculate the terminal depreciation for
        # situations.where there is terminal growth. This will enforce that
        # Capex>=Depreciation so that assets continue to increase as the
        # company grows its bottom line

        C = self.gt/self.roict*(1-self.t)
        dat = (self.fin['capex'].iloc[-1]-C*self.fin['ebitda'].iloc[-1])/(1-C)
        if dat < 0:
            logging.error('negative depreciation in terminal year, check roic and growth assumptions')

        self.fin['da'] = self.__stream(self.fin['da'], dat)
        self.fin['income_pretax'] = self.fin.ebitda - \
            self.fin.sbc - self.fin.da - self.fin.interest
        self.fin['dDebt'] = self.fin['debt']-self.fin['debt'].shift(1)
        # todo: calculate from interest0 and Debt0
        self.fin['dDebt'].iloc[0] = 0
        
        #Calculating taxes

        self.fin['tax_cash'] = np.nan
        self.fin['tax_cash'].iloc[0] = self.fin['tax'].iloc[0]
  
        if self.te is None:
            tax0 = max(self.t*(self.fin['income_pretax'].iloc[0]),self.fin['tax'].iloc[0]) #important to avoid -'ve taxes with NOL's, also so they don't have unsustainably low taxes
            tax1 = tax0 + self.t * (self.fin['income_pretax'].iloc[1] - self.fin['income_pretax'].iloc[0]) 
            te = tax1/self.fin['income_pretax'].iloc[1]
        else:
            te = self.te #uses the value set during initialization
        
        self.fin['income_taxable'] = np.nan
        self.fin['income_taxable'].iloc[0] = max(self.fin['income_pretax'].iloc[0]*(1-self.fin['nol'].iloc[0] > 0), 0) #zero the income in the baseline year if NOL>0
        for i in range(1, self.year+1):
            self.fin['nol'].iloc[i] = max(self.fin['nol'].iloc[i-1] - self.fin['income_pretax'].iloc[i], 0)
            self.fin['income_taxable'].iloc[i] = max(0, self.fin['income_pretax'].iloc[i] - self.fin['nol'].iloc[i-1])
            if i == 1:
                self.fin['tax'].iloc[1] = te*max(self.fin['income_pretax'].iloc[1],0)
                self.fin['tax_cash'].iloc[i] = self.fin['tax'].iloc[1] + self.t * min((self.fin['nol'].iloc[i] - self.fin['nol'].iloc[i-1]),0)
            else:
                self.fin['tax'].iloc[i] = self.fin['tax'].iloc[i-1]+self.t * (max(self.fin['income_pretax'].iloc[i],0) - max(self.fin['income_pretax'].iloc[i-1],0))
                self.fin['tax_cash'].iloc[i] = self.fin['tax'].iloc[i]+self.t * min((self.fin['nol'].iloc[i] - self.fin['nol'].iloc[i-1]),0)
                     
        #calculating FCF

        self.fin['fcf'] = self.fin.ebitda - self.fin.sbc - \
            self.fin.tax_cash - self.fin.capex - self.fin.dwc - self.fin.interest
        self.fin['fcfe'] = self.fin.ebitda - self.fin.sbc - self.fin.tax_cash - \
            self.fin.capex - self.fin.dwc + self.fin.dDebt - self.fin.interest - self.fin.MnA
        self.fin['fcff'] = self.fin.ebitda - self.fin.sbc - self.fin.tax_cash - \
            self.fin.capex - self.fin.dwc - self.fin.interest*self.t - self.fin.MnA

        self.fin['dividend_policy'] = 0
        n_div = len(self.dividend)
        for i in range(self.year+1):
            if (i < n_div):
                self.fin['dividend_policy'].iloc[i] = self.dividend[i] * \
                    self.fin['shares'].iloc[i]
            else:
                self.fin['dividend_policy'].iloc[i] = max(self.dividend[n_div-1]*self.fin['shares'].iloc[n_div-1] /
                                                          self.fin['fcf'].iloc[n_div-1]*self.fin['fcf'].iloc[i], self.fin['dividend_policy'].iloc[i-1])
        self.fin['dividend'] = (self.fin['fcfe']-self.fin['buybacks'])/self.fin['shares']
        self.fin['cash'].iloc[1:] = self.fin['fcfe'].iloc[1:]
        self.fin['cash'] = self.fin['cash'].cumsum()
        self.fin['noa'].iloc[1:] = self.fin['noa'].iloc[0]
        logging.info('fcf_from_ebitda() method complete')

    def fcf_to_debt(self, leverage=3, year_d=1):
        '''Adjust debt levels to desired target.

        Decrease (or increase) FCFE to reduce (or increase) debt towards target
        leverage prerequisite: Must first have fcf defined

        Args:
            leverage = desired Debt/EBITDA, default value is 3
        Returns:
        '''

        # increase debt if fcf is negative and cash is 0
        if self.data_for_ebitda is False:
            logging.error('financial dataset cannot be used to optimize leverage')

        if self.fin['fcf'].empty:
            logging.error('first calculate fcf')

        self.fin['debt_Target'] = [leverage*self.fin['ebitda'].iloc[i]
                                   if self.fin['ebitda'].iloc[i] > 0 else 0 for i in range(self.year+1)]

        # run the loop 3 times just to converge on the interest and FCF
        for i in range(3):
            for i in range(year_d-1, self.year):
                if (self.fin['debt'].iloc[i]-self.fin['debt_Target'].iloc[i+1] < 0):  # underlevered
                    dDebt = -1*(self.fin['debt'].iloc[i] -
                                self.fin['debt_Target'].iloc[i+1])
                else:  # overlevered
                    if i == 0:
                        dDebt = -1*min(self.fin['debt'].iloc[i]-self.fin['debt_Target'].iloc[i+1], self.fin['fcf'].iloc[i+1] +
                                       self.fin['cash'].iloc[i]-self.fin['MnA'].iloc[i+1]-self.fin['dividend_policy'].iloc[i+1])
                    else:
                        dDebt = -1*min(self.fin['debt'].iloc[i]-self.fin['debt_Target'].iloc[i+1],
                                       self.fin['fcf'].iloc[i+1]-self.fin['MnA'].iloc[i+1]-self.fin['dividend_policy'].iloc[i+1])
                self.fin['debt'].iloc[i+1] = self.fin['debt'].iloc[i]+dDebt
            self.fcf_from_ebitda()
        logging.info('fcf_to_debt() method complete')

    def fcf_to_bs(self):
        '''
        Args:

        Returns:
        '''
        self.fin['cashBS'].iloc[0] = self.fin['cash'].iloc[0]
        for i in range(self.year):
            self.fin['cashBS'].iloc[i+1] = self.fin['cashBS'].iloc[i] + self.fin['fcfe'].iloc[i +
                                                                                              1] - self.fin['dividend_policy'].iloc[i+1] - self.fin['buybacks'].iloc[i+1]

        self.fin['dividend'] = self.fin['dividend_policy']/self.fin['shares']
        self.fin['dividend'].iloc[-1] = self.fin['dividend'].iloc[-1] + self.fin['cashBS'].iloc[-1] / \
            self.fin['shares'].iloc[-1]  # all remaining cash distributed the year before terminal
        self.cash0 = 0  # discount future cash back to NPV
        logging.info('fcf_to_bs() method complete')

    def fcf_to_buyback(self, price, dp='proportional'):
        '''Use cash balance to buyback shares and reduce sharecounts

        prerequisite: Must first have fcf defined

        Args:
            price: share price; could be todays shareprice or anything else
            dp: method for share price change, 'constant' or 'proportional',
            'constant' = maintain constant share price,
            'proportional' = constant EV/EBITDA to todays value

        Returns:
        '''
        self.fin['price'] = price
        # limit buybacks to when fcf>0
        if dp == 'constant':
            for i in range(self.year):
                if i == 0:
                    self.fin['buybacks'].iloc[i+1] = self.fin['fcfe'].iloc[i+1] + \
                        self.fin['cash'].iloc[0] - \
                        self.fin['dividend_policy'].iloc[i+1]
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - \
                        self.fin['buybacks'].iloc[i+1] / \
                        self.fin['price'].iloc[i]
                else:
                    self.fin['buybacks'].iloc[i+1] = self.fin['fcfe'].iloc[i +
                                                                           1] - self.fin['dividend_policy'].iloc[i+1]
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - \
                        self.fin['buybacks'].iloc[i+1] / \
                        self.fin['price'].iloc[i]
        elif dp == 'proportional':
            EV = price*self.shares + \
                self.fin['debt'].iloc[0]-self.fin['cash'].iloc[0]
            # calculate the forward multiple
            multiple = EV/self.fin['ebitda'].iloc[1]
            for i in range(self.year-1):
                if i == 0:
                    self.fin['buybacks'].iloc[i+1] = self.fin['fcfe'].iloc[i+1] + \
                        self.fin['cash'].iloc[0] - \
                        self.fin['dividend_policy'].iloc[i+1]
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - \
                        self.fin['buybacks'].iloc[i+1] / \
                        self.fin['price'].iloc[i]
                else:
                    self.fin['buybacks'].iloc[i+1] = self.fin['fcfe'].iloc[i +
                                                                           1] - self.fin['dividend_policy'].iloc[i+1]
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - \
                        self.fin['buybacks'].iloc[i+1] / \
                        self.fin['price'].iloc[i]
                # calculate the new price
                # no need to include cash, since cash is being used fully for buybacks or dividend
                self.fin['price'].iloc[i+1] = max((multiple*self.fin['ebitda'].iloc[i+2] - self.fin['debt'].iloc[i+1])/self.fin['shares'].iloc[i+1],self.fin['price'].iloc[i])
                
            self.fin['price'].iloc[-1] = self.fin['price'].iloc[-2]
            self.fin['shares'].iloc[-1] = self.fin['shares'].iloc[-2] - self.fin['buybacks'].iloc[-1] / self.fin['price'].iloc[-1]

        self.fin['dividend'] = (
            self.fin['fcfe']-self.fin['buybacks'])/self.fin['shares']
        self.fin['dividend'].iloc[0] = self.dividend[0]
        self.fin['dividend'].iloc[1] = (
            self.fin['fcfe'].iloc[1]+self.cash0-self.fin['buybacks'].iloc[1])/self.fin['shares'].iloc[1]
        self.cash0 = 0  # all used for buybacks, you need to zero it so that it's not double counted in the valuation for the DDM model
        self.buybacks = True
        logging.info('fcf_to_buyback() method complete')

    def fcf_to_allocate(self, price, dp='proportional', buybacks=None):
        '''A generalized method for allocating cash to dividends, buybacks or
        storing on the balance sheet

        prerequisite: Must first have fcf defined

        Args:
            price: share price; could be todays shareprice or anything else
            dp: method for share price change, 'constant' or 'proportional',
            'constant' = maintain constant share price,
            'proportional' = constant EV/EBITDA to todays value
            buybacks: (counter-intuitvly) 'None' means as much as possible and
            is the default. Otherwise input an array with the pedefined amount.
            The remaing cash is stored on the balance sheet.

        Returns:

        '''
        if buybacks is None:
            pass
        elif isinstance(buybacks, list):
            self.buybacks = buybacks
        elif isinstance(buybacks, float):
            self.buybacks = [buybacks]
        elif isinstance(buybacks, int):
            self.buybacks = [buybacks]

        # set buyback level
        if buybacks is None:  # all FCF not used for dividends are used for BB's
            self.fcf_to_buyback(price, dp)
        else:  # set a specific BB level and accumulate the remaing cash onto the BS
            self.fin['buybacks'] = 0
            n_bb = len(self.buybacks)
            for i in range(self.year+1):
                if (i < n_bb):
                    self.fin['buybacks'].iloc[i] = self.buybacks[i]
                else:
                    self.fin['buybacks'].iloc[i] = self.buybacks[n_bb-1] / \
                        self.fin['fcf'].iloc[n_bb-1]*self.fin['fcf'].iloc[i]

            # calculate price and shares
            self.fin['price'] = price
            if dp == 'constant':
                for i in range(self.year):
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - \
                        self.fin['buybacks'].iloc[i+1] / \
                        self.fin['price'].iloc[i]
            elif dp == 'proportional':
                EV = price*self.shares + \
                    self.fin['debt'].iloc[0]-self.fin['cash'].iloc[0]
                # calculate the forward multiple
                multiple = EV/self.fin['ebitda'].iloc[1]
                for i in range(self.year-1):
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - \
                        self.fin['buybacks'].iloc[i+1] / \
                        self.fin['price'].iloc[i]
                    # no need to include cash, since cash is being used fully for buybacks or dividend
                    self.fin['price'].iloc[i+1] = max((multiple*self.fin['ebitda'].iloc[i+2] - self.fin['debt'].iloc[i+1])/self.fin['shares'].iloc[i+1],self.fin['price'].iloc[i])
                #self.fin['shares'].iloc[-1] = self.fin['shares'].iloc[-2]
                self.fin['price'].iloc[-1] = self.fin['price'].iloc[-2]
                self.fin['shares'].iloc[-1] = self.fin['shares'].iloc[-2] - self.fin['buybacks'].iloc[-1] / self.fin['price'].iloc[-1]

        self.fcf_to_bs()
        self.buybacks = True

    def fcf_to_acquire(self, adjust_cash, year_a=1, ebitda_frac=0.1, multiple=10, leverage=3, gnext=0.1, cap_frac=0.2):
        '''include the effect of an acquistion in the financial model

        Models the effect of an acquisition.
        Increases EBITDA according to multiple paid.
        Capital structure of the target is specified so that debt is adjusted
        accordingly.
        Cash outlay in excess of debt is included in the MnA key of the fin
        dataframe to reduce FCFE/FCFF accordingly.

        Args:
            ebitda_frac: EBITDA of the target, relative to the organic ebitda
            gnext: next years growth
            multiple: EV/EBITDA multiple of the acquisition
            leverage: Debt/EBITDA leverage target
            adjust_cash: adjust cash balance in year 0

        Returns:
            dEbitda: an array showing the change in EBITDA as result of the
            acquisition

        '''

        if self.data_for_ebitda is False:
            logging.error('financial dataset cannot be used to acquire')
        if self.fin['fcf'].empty:
            logging.error('first calculate fcf')

        g = [ebitda_frac-1, gnext]
        for i in range(year_a):
            g.insert(0, 0)
        dEbitda = self.forecast_ebitda(self.fin['ebitda'].iloc[year_a], g)
        for i in range(year_a+1):
            dEbitda[i] = 0

        dCapex = cap_frac*np.array(dEbitda)
        dDebt = [leverage*dEbitda[year_a+1] if x >=
                 year_a else 0 for x in range(self.year+1)]
        self.fin['debt'] = self.fin['debt']+dDebt
        self.fin['MnA'].iloc[year_a] = self.fin['MnA'].iloc[year_a] + \
            multiple*dEbitda[year_a+1]
        self.fin['capex'] = self.fin['capex']+dCapex
        self.fin['ebitda'] = self.fin['ebitda']+dEbitda
        # reset the depreciation so that it gets recalculated from fcf_from_ebitda
        self.fin['da'].iloc[year_a+1:] = np.nan

        if year_a == 0 and adjust_cash is True:
            # adjust the cash balance in year 0 to pay for the acquisition
            self.fin['cash'].iloc[0] = self.fin['cash'].iloc[0] - \
                (multiple-leverage)*dEbitda[1]
            self.cash0 = self.fin['cash'].iloc[0]

        if self.fin['cash'].iloc[year_a] < 0:
            logging.error('cash<0, insufficient cash for the aquisition; lower the EBITDA or increase the leverage')

        self.fcf_from_ebitda()
        logging.info('fcf_to_acquire() method complete')

        return dEbitda

    def noa_to_dispose(self, dnoa, tax=0, year_dis=1):
        '''dispose of non-operating assets (noa). The effect of this is to
        reduce the amount of 'noa' and create a (negative) MnA entry in the
        'fin' dataframe

        Args:
            dnoa: gross amount of the transaction
            tax: tax rate (a fraction)
            year_dis: year in which it is disposed

        Returns:
        '''
        self.fin['MnA'].iloc[year_dis] = self.fin['MnA'].iloc[year_dis] - \
            dnoa*(1-tax)
        self.fin['noa'] = self.fin['noa'] - dnoa
        self.fcf_from_ebitda()
        logging.info('dispose_from_noa() method complete')

    def value(self):
        '''calculate the firm and equity values

        Firm value is calculated as the DCF of the FCFF. Equity value is
        calculated using the DCF of the FCFE. Value per share is also
        calculated by dividing by the current sharecount. Effects of capital
        allocation decisions such as dividends and buybacks can be factored
        into the analysis by looking at the dividend discount model (DDM).
        Results from the analysis are used to populate the 'fin' dataframe.

        Args:

        Returns:
            self.fin['equity']: DCF of the FCFE
            self.fin['firm']: DCF of the FCFF

        '''

        if self.data_for_ebitda is True:
            # really complicated way to calculate the terminal FCFE for situtions...
            # ...where there is terminal growth and you have changes in debt the final year before terminal.
            # If no change in debt and no growth the fcfet = fcfe = fcf in the terminal year
            self.fcfet = (self.fin['fcf'].iloc[-1]-self.fin['dDebt'].iloc[-1]
                          * self.rd+self.fin['debt'].iloc[-1]*self.gt)*(1+self.gt)

            self.fin['equity'] = self.__pv(
                cfs=self.fin.fcfe, cft=self.fcfet, g=self.gt, r=self.re)
            self.__wacc()
            self.fin['firm'] = self.__pv(
                cfs=self.fin.fcff, g=self.gt, r=self.fin.wacc)
            self.fin['DDM'] = self.__pv(
                cfs=self.fin.dividend, cft=self.fcfet/self.fin['shares'].iloc[-1], g=self.gt, r=self.re)

            # adjustments for cash and non-operating assets
            self.fin['value_per_share'] = (
                self.fin['equity']+self.fin['noa'])/self.shares
            self.fin['value_per_share'].iloc[0] = self.fin['value_per_share'].iloc[0] + \
                self.fin['cash'].iloc[0]/self.shares
            self.fin['value_per_share_DDM'] = self.fin['DDM'] + \
                self.fin['noa']/self.fin['shares']
            self.fin['value_per_share_DDM'].iloc[0] = self.fin['value_per_share_DDM'].iloc[0] + \
                self.cash0/self.fin['shares'].iloc[0]
            self.fin['value_per_share_DDM'].iloc[-1] = self.fin['value_per_share_DDM'].iloc[-1] + \
                self.fin['cashBS'].iloc[-1]/self.fin['shares'].iloc[-1]
        else:
            self.fin['equity'] = self.__pv(
                cfs=self.fin.fcfe, g=self.gt, r=self.re)
            self.fin['firm'] = None

        if self.buybacks is True:
            # calculate value per share in buyback scenario, i.e. all cash used
            # to purchase shares until terminal year
            self.vpsbb = ((self.fin['equity'].iloc[-1]+self.fin['noa'].iloc[-1]
                           )/self.fin['shares'].iloc[-1])/(1+self.re)**self.year
        else:
            self.vpsbb = 0

        logging.info('value() method complete')
        return self.fin['equity'], self.fin['firm']

    def display_fin(self):
        '''populates a copy of the excel template file with a summary of the
        financial analysis contained in the fin dataframe.

        Args:

        Returns:
            table: a summary table of the financial analysis
        '''
        table = self.fin[['revenue','ebitda', 'sbc', 'da', 'interest', 'income_pretax', 'nol', 'income_taxable', 'tax_cash', 'tax', 'capex', 'MnA', 'dDebt', 'dwc', 'fcf', 'fcfe', 'fcff',
                          'buybacks', 'dividend', 'cash', 'cashBS', 'noa', 'equity', 'debt', 'EV', 'wacc', 'firm', 'shares', 'price', 'value_per_share', 'value_per_share_DDM']].T.style.format("{:.1f}")

        wb = load_workbook(filename=os.path.join(
            os.path.dirname(__file__), '..\\')+'company_template.xlsx')
        ws = wb['raw data']

        for r in dataframe_to_rows(self.fin.T, index=True, header=True):
            ws.append(r)
        ws['A1'] = 'date'

        ws = wb['report']
        ws['B2'] = self.ticker
        ws['B3'] = self.now
        ws['B4'] = 'David May'

        ws['B6'] = self.rd
        ws['B7'] = self.re
        ws['B8'] = self.gt
        ws['B9'] = self.t

        ws['B11'] = self.fin['cash'].iloc[-1]
        ws['B12'] = self.fcfet
        ws['B13'] = self.fin['equity'].iloc[-1]
        ws['B15'] = self.shares
        ws['B17'] = self.fin['value_per_share_DDM'].iloc[0]

        ws2 = wb['report']

        wb.save(self.ticker+'.xlsx')
        return table
